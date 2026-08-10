"""
Ritvaa Sales Dashboard builder
================================
Turns your "Sales Detail" Excel export into a self-contained, interactive
HTML dashboard (KPIs + charts + table). No internet connection needed to
view the result - Chart.js and its data-labels plugin are bundled
directly into the output file.

HOW TO USE
----------
1. Edit the EXCEL_PATH line below to point at your .xlsx file.
   (Or, if you prefer, just run:  python3 build_dashboard.py "C:/path/to/your/file.xlsx")
2. Run this script:   python3 build_dashboard.py
3. Open the file it creates (dashboard.html) in your browser. Done.

Keep this script in the same folder as:
  - dashboard_template.html    (the dashboard's HTML/CSS/JS shell)
  - chart.umd.js               (the bundled charting library)
  - datalabels.min.js          (bundled plugin that draws numbers on bars)
These are used only while generating the file - the final dashboard.html
does not need them anymore, it's fully self-contained.
"""

import sys
import json
import os
from datetime import datetime, date

# ============================================================
# 1) SET THE PATH TO YOUR EXCEL FILE HERE
# ============================================================
EXCEL_PATH = "Order Data.xlsx"   # <-- CHANGE THIS to your real file's path
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "dashboard_template.html")
CHART_JS_PATH = os.path.join(SCRIPT_DIR, "chart.umd.js")
DATALABELS_JS_PATH = os.path.join(SCRIPT_DIR, "datalabels.min.js")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "dashboard.html")


def get(row, *keys, default=""):
    for k in keys:
        if k in row and row[k] not in (None, ""):
            return row[k]
    return default


def to_number(v):
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace("₹", "").replace(",", "").strip())
    except ValueError:
        return 0


def to_iso_date(v):
    if isinstance(v, (datetime, date)):
        if isinstance(v, date) and not isinstance(v, datetime):
            v = datetime.combine(v, datetime.min.time())
        return v.isoformat()
    if v in (None, ""):
        return datetime.now().isoformat()
    s = str(v).strip()
    # Try the common Excel export format: "01/08/2026 07:22 AM"
    for fmt in ("%d/%m/%Y %I:%M %p", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).isoformat()
        except ValueError:
            continue
    return datetime.now().isoformat()


def normalize_row(row):
    order_no = get(row, "Order No", "orderNo")
    if not order_no:
        return None
    return {
        "orderNo": str(order_no),
        "externalOrderNo": str(get(row, "External Order No", "externalOrderNo")),
        "orderDate": to_iso_date(get(row, "Order Date", "orderDate")),
        "orderType": get(row, "Order Type", "orderType", default="Prepaid"),
        "status": get(row, "Status", "status", default="Processing"),
        "country": get(row, "Country", "country", default="India"),
        "state": get(row, "State", "state", default="Unknown"),
        "city": get(row, "City", "city", default="Unknown"),
        "skuCode": get(row, "SKU Code", "skuCode"),
        "skuDesc": get(row, "SKU Desc", "skuDesc", default="Item"),
        "category": get(row, "Category1", "category", default="Uncategorized"),
        "subCategory": get(row, "Sub Category", "subCategory"),
        "size": str(get(row, "Size", "size")),
        "quantity": to_number(get(row, "Quantity", "quantity")) or 1,
        "price": to_number(get(row, "Price", "price")),
        "shipCost": to_number(get(row, "Ship Cost", "shipCost")),
        "packingCost": to_number(get(row, "Packing Cost", "packingCost")),
        "mrp": to_number(get(row, "MRP", "mrp")),
        "discount": to_number(get(row, "Discount", "discount")),
        "discountCode": get(row, "Discount Code", "discountCode"),
        "tax": to_number(get(row, "Tax", "tax")),
        "invoiced": to_number(get(row, "Invoiced", "invoiced")),
        "cogs": to_number(get(row, "COGS", "cogs")),
        "grossMargin": to_number(get(row, "Gross Margin", "grossMargin")),
        "gmPercent": to_number(get(row, "GM Percent", "gmPercent")),
        "onHoldStatus": get(row, "On Hold Status", "onHoldStatus", default="No"),
        "replacementOrder": get(row, "Replacement Order", "replacementOrder", default="No"),
        "vendorName": get(row, "VendorName", "vendorName", default="Unassigned"),
        "channelCode": get(row, "Channel Code", "channelCode"),
        "channelName": get(row, "Channel Name", "channelName", default="Unknown"),
        "customerCode": str(get(row, "customerCode", "Email", "email", default="unknown")),
    }


def load_excel(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit("Missing dependency. Run:  pip install openpyxl --break-system-packages")

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == "sales detail":
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit(f"Sheet '{sheet_name}' is empty.")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    records = []
    for raw in rows[1:]:
        if all(v is None or v == "" for v in raw):
            continue
        row = {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}
        normalized = normalize_row(row)
        if normalized:
            records.append(normalized)
    return records, sheet_name


def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else EXCEL_PATH

    if not os.path.exists(excel_path):
        sys.exit(
            f'Could not find "{excel_path}".\n'
            f"Edit EXCEL_PATH at the top of build_dashboard.py, or run:\n"
            f'  python3 build_dashboard.py "C:/path/to/your/file.xlsx"'
        )

    print(f"Reading {excel_path} ...")
    records, sheet_name = load_excel(excel_path)
    if not records:
        sys.exit(f'No usable rows found in sheet "{sheet_name}" (need an "Order No" column).')
    print(f"  -> {len(records)} orders loaded from sheet '{sheet_name}'")

    dates = sorted(r["orderDate"] for r in records)
    print(f"  -> date range: {dates[0][:10]} to {dates[-1][:10]}")

    for label, path in (("dashboard_template.html", TEMPLATE_PATH),
                         ("chart.umd.js", CHART_JS_PATH),
                         ("datalabels.min.js", DATALABELS_JS_PATH)):
        if not os.path.exists(path):
            sys.exit(f"Missing {path} - keep this script next to {label}")

    template = open(TEMPLATE_PATH, encoding="utf-8").read()
    chart_js = open(CHART_JS_PATH, encoding="utf-8").read()
    datalabels_js = open(DATALABELS_JS_PATH, encoding="utf-8").read()

    html = template.replace("__CHART_JS__", chart_js)
    html = html.replace("__DATALABELS_JS__", datalabels_js)
    html = html.replace("__SALES_DATA__", json.dumps(records))
    html = html.replace(
        "__DATA_SOURCE_LABEL__",
        f"{len(records)} orders from {os.path.basename(excel_path)}",
    )

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"Done. Open {OUTPUT_PATH} in your browser.")


if __name__ == "__main__":
    main()
